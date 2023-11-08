import re

import pandas # "./test_data/DataSet.xlsx"
import openpyxl
import unittest


import html_processing
import nlp_processing

class TestQuantitativeResults(unittest.TestCase):

    def test_old_new(self, url_old=None, url_new=None, result=None):
        if url_old is None or url_new is None or result is None:
            print("EMPTY TEST")
            pass
        else:
            if re.match("C:\s\d+\s([+]\s\d+\s)*; M:\s\d+\s([+]\s\d+\s)*;", result):
                "Extracting the data from the result column to integer usable for comparison"
                expected_cs = result.split(";")[0].replace("C: ", "").split("+")
                expected_cs_changes = len(expected_cs)
                expected_cs_mods_per_changes = []
                for ecs in expected_cs:
                    expected_cs_mods_per_changes.append(int(ecs))
                    if int(ecs) == 0:
                        expected_cs_changes -= 1
                expected_ms = result.split(";")[1].replace("M: ", "").split("+")
                expected_ms_changes = len(expected_ms)
                expected_ms_mods_per_changes = []
                for ems in expected_ms:
                    expected_ms_mods_per_changes.append(int(ems))
                    if int(ems) == 0:
                        expected_ms_changes -= 1

                overall_expected_mods = 0
                for modcs in expected_cs_mods_per_changes:
                    overall_expected_mods += modcs
                for modms in expected_ms_mods_per_changes:
                    overall_expected_mods += modms

                print(f"C: {expected_cs_changes} and Modifications per Changes = {expected_cs_mods_per_changes}")
                print(f"M: {expected_ms_changes} and Modifications per Changes = {expected_ms_mods_per_changes}")
                print(f"Overall there are {overall_expected_mods} expected!")

                "Run the algorithm (either acurate or fast)"
                celex_old, doc_old = html_processing.pars_html(url_old)
                celex_new, doc_new = html_processing.pars_html(url_new)
                html_result = html_processing.find_changes_and_make_diff_of_surrounding_text(doc_old, doc_new)
                nlp_result = nlp_processing.process_nlp("testfast" + celex_old + celex_new, html_result, True)  # fast test
                #nlp_result = nlp_processing.process_changes("testaccurate" + celex_old + celex_new, html_result,False)  # accurate test todo

                if "repealed" in result:
                    print(nlp_result is None)
                    self.assertEqual(nlp_result, None) # if repealed
                elif expected_ms_changes + expected_cs_changes == 0:
                    print(nlp_result is None)
                    self.assertEqual(nlp_result, None) # if there are no changes
                else:
                    found_ms_changes = 0
                    found_cs_changes = 0
                    for change_name in nlp_result[0]:
                        if change_name.startswith("A") or change_name.startswith("M"):
                            found_ms_changes += 1
                        if change_name.startswith("C"):
                            found_cs_changes += 1
                    overall_found_mods = 0
                    for frs in nlp_result[1]:
                        overall_found_mods += len(frs)
                    overall_found_mods -= len(nlp_result[1])
                    print(f"Test 1: found {len(nlp_result[1])}, expected {expected_ms_changes + expected_cs_changes}")
                    print(len(nlp_result[1]) == expected_ms_changes + expected_cs_changes)
                    self.assertEqual(len(nlp_result[1]), expected_ms_changes + expected_cs_changes)  # test 1: amount of changes
                    print(len(nlp_result[0]) == expected_ms_changes + expected_cs_changes)
                    self.assertEqual(len(nlp_result[0]), expected_ms_changes + expected_cs_changes)  # test 1: amount of changes
                    print(f"Test 2: found Cs {found_cs_changes} expected {expected_cs_changes}; found Ms {found_ms_changes}, expected {expected_ms_changes}")
                    print(found_ms_changes == expected_ms_changes)
                    self.assertEqual(found_ms_changes, expected_ms_changes)  # test 2: amount of amendments
                    print(found_cs_changes == expected_cs_changes)
                    self.assertEqual(found_cs_changes, expected_cs_changes)  # test 2: amount of corridgendi
                    print(f"Test 3: found {overall_found_mods}, expected {overall_expected_mods}")
                    print(overall_found_mods == overall_expected_mods)
                    self.assertEqual(overall_found_mods, overall_expected_mods) # test 3: amount of all modifications
                    return overall_found_mods, overall_expected_mods
            else:
                self.assertEqual(0,13)
                print("RESULT not usable")
        return 0,0

    def test_file(self):
        dataframe = openpyxl.load_workbook("./test_data/DataSetWithFullResults.xlsx")
        data = dataframe.active

        expected_mod_dir = [0] * 20
        found_mod_dir = [0] * 20

        for row in range(2, data.max_row):
            print("ROW -----------------------------------")
            print(row)
            # [celex, language, celex_link, first_date, first_url, middle_date, middle_url, last_date, last_url, amount_of_consolidated_versions, directory, topic, name, comment, ...
            # ..., retrieval_date, result_first_last, result_middle_last, result_first_middle]
            row_entries = list(data.iter_cols(1, data.max_column))
            celex = row_entries[1][row].value.__str__()
            print("celex: " + celex)
            language = row_entries[2][row].value.__str__()
            # print("language: " + language)
            celex_link = row_entries[3][row].value.__str__()
            # print("celex_link: " + celex_link)
            first_date = row_entries[4][row].value.__str__()
            # print("first_date: " + first_date)
            first_url = row_entries[5][row].value.__str__()
            # print("first_url: " + first_url)
            middle_date = row_entries[6][row].value.__str__()
            # print("middle_date: " + middle_date)
            middle_url = row_entries[7][row].value.__str__()
            # print("middle_url: " + middle_url)
            last_date = row_entries[8][row].value.__str__()
            # print("last_date: " + last_date)
            last_url = row_entries[9][row].value.__str__()
            # print("last_url: " + last_url)
            amount_of_consolidated_versions = row_entries[10][row].value.__str__()
            # print("amount_of_consolidated_versions: " + amount_of_consolidated_versions)
            directory = row_entries[11][row].value.__str__()
            # print("directory: " + directory)
            topic = row_entries[12][row].value.__str__()
            # print("topic: " + topic)
            name = row_entries[13][row].value.__str__()
            # print("name: " + name)
            comment = row_entries[14][row].value.__str__()
            print("comment: " + comment)
            retrieval_date = row_entries[15][row].value.__str__()
            # print("retrieval_date: " + retrieval_date)
            result_first_last = row_entries[16][row].value.__str__()
            print("result_first_last: " + result_first_last)
            result_middle_last = row_entries[17][row].value.__str__()
            # print("result_middle_last: " + result_middle_last)
            result_first_middle = row_entries[18][row].value.__str__()
            # print("result_first_middle: " + result_first_middle)

            if row < 0: # 0, 101, 201
                continue
            elif row > 300: # 100, 200, 300
                break
            else:
                if (first_url is not None and "NULL" not in first_url and "Not" not in first_url
                        and last_url is not None and "NULL" not in last_url and "Not" not in last_url
                        and result_first_last is not None and "OVERFLOW" not in result_first_last and "NULL" not in result_first_last):
                    with self.subTest(celex + first_date + last_date + result_first_last):
                        flf, fle = self.test_old_new(first_url, last_url, result_first_last)
                        "To add up the modifications (comment the assertion test!)"
                        '''if directory is not None:
                            print(f"To Directory {directory} the expected {fle} and found {flf} are added")
                            try:
                                directory = int(directory)
                                if 0 < directory < 21:
                                    found_mod_dir[directory - 1] += flf
                                    expected_mod_dir[directory - 1] += fle
                                    print(found_mod_dir)
                                    print(expected_mod_dir)
                            except:
                                print("Calc-Error")'''
                if (middle_url is not None and "NULL" not in middle_url and "Not" not in middle_url
                        and last_url is not None and "NULL" not in last_url and "Not" not in last_url
                        and result_middle_last is not None and "OVERFLOW" not in result_middle_last and "NULL" not in result_middle_last):
                    with self.subTest(celex + middle_date + last_date + result_middle_last):
                        self.test_old_new(middle_url, last_url, result_middle_last)
                if (first_url is not None and "NULL" not in first_url and "Not" not in first_url
                    and middle_url is not None and "NULL" not in middle_url and "Not" not in middle_url
                    and result_first_middle is not None and "OVERFLOW" not in result_first_middle and "NULL" not in result_first_middle):
                    with self.subTest(celex + first_date + middle_date + result_first_middle):
                        self.test_old_new(first_url, middle_url, result_first_middle)
        "The result is only accurate, if the self.assert() tests are commented out, because failing tests return before the value is return!"
        '''print("FINAL RESULT: (mod)")
        print("Found")
        print(found_mod_dir)
        print("Expected")
        print(expected_mod_dir)'''
        return